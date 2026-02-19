from __future__ import annotations

from pathlib import Path
import csv
import re
from typing import Dict, Iterator, List, Optional, Tuple, Set

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import column_index_from_string


PLANTILLA = Path(__file__).parent / "RIPS COMFE_PLANTILLA.xlsm"
HOJA_ESTRUCTURA = "ESTRUCTURA"
HOJA_US = "US"

CONTROL_SHEET = "__RIPS_CONTROL__"  # hoja oculta para persistir DOCs de US

# Mapeos ya validados (CSV index -> Columna Excel)
MAPEO: Dict[str, Dict[int, str]] = {
    "AT": {3: "E", 4: "F", 7: "K", 11: "L"},
    "AP": {3: "E", 4: "F", 10: "I", 15: "K", 16: "L"},
    "AC": {3: "E", 4: "F", 9: "I", 17: "K", 18: "L"},
}

# Precompilado: CSV index -> Excel column index (int)
_MAPEO_IDX: Dict[str, List[Tuple[int, int]]] = {
    k: [(csv_i, column_index_from_string(col)) for csv_i, col in m.items()]
    for k, m in MAPEO.items()
}

_re_non_digits = re.compile(r"\D+")


def _norm_doc(v) -> str:
    """Normaliza documento para US: solo dígitos, sin puntos/guiones/espacios."""
    if v is None:
        return ""
    s = str(v).strip()
    return _re_non_digits.sub("", s)


def iterar_csv_sin_header(path: Path) -> Iterator[List[str]]:
    """Itera un CSV saltando header sin cargarlo completo a memoria."""
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        r = csv.reader(f)
        next(r, None)  # header
        for row in r:
            yield row


def _siguiente_fila_por_columna(
    ws: Worksheet,
    col_letter: str,
    fila_min: int = 2,
    max_scan: int = 200000,
) -> int:
    """
    Retorna la siguiente fila libre (1 + última fila con valor en col_letter).
    Evita depender de ws.max_row cuando hay formatos/estilos extendidos.
    """
    col_idx = column_index_from_string(col_letter)
    end = ws.max_row
    start = max(fila_min, end - max_scan)

    for row_i in range(end, start - 1, -1):
        v = ws.cell(row=row_i, column=col_idx).value
        if v not in (None, ""):
            return row_i + 1

    return fila_min


def _get_or_create_control_sheet(wb: Workbook) -> Worksheet:
    if CONTROL_SHEET in wb.sheetnames:
        ws = wb[CONTROL_SHEET]
    else:
        ws = wb.create_sheet(CONTROL_SHEET)
        ws["A1"].value = "KIND"   # Solo usaremos "U"
        ws["B1"].value = "KEY"    # DOC normalizado
        ws.sheet_state = "hidden"
    ws.sheet_state = "hidden"
    return ws


def _load_us_set(ws_control: Worksheet) -> Set[str]:
    seen_u: Set[str] = set()
    for kind, key in ws_control.iter_rows(min_row=2, max_col=2, values_only=True):
        if kind == "U" and key:
            seen_u.add(str(key))
    return seen_u


def _append_control(ws_control: Worksheet, kind: str, key: str) -> None:
    r = ws_control.max_row + 1
    ws_control.cell(row=r, column=1, value=kind)
    ws_control.cell(row=r, column=2, value=key)


def abrir_plantilla(
    path: Optional[Path] = None,
    *,
    keep_vba: bool = True,
) -> Tuple[Workbook, Worksheet, Worksheet, Worksheet]:
    """Abre la plantilla UNA vez y retorna (wb, ws_estructura, ws_us, ws_control)."""
    p = Path(path) if path else PLANTILLA
    wb = load_workbook(p, keep_vba=keep_vba)
    ws = wb[HOJA_ESTRUCTURA]
    ws_us = wb[HOJA_US]
    ws_control = _get_or_create_control_sheet(wb)
    return wb, ws, ws_us, ws_control


def cargar_en_hojas(
    carpeta_csv: Path,
    ws_estructura: Worksheet,
    ws_us: Worksheet,
    ws_control: Worksheet,
    fila_estructura: Optional[int] = None,
    fila_us: Optional[int] = None,
    *,
    imprimir_cada: int = 1000,
) -> Tuple[int, int]:
    """
    Carga AT/AP/AC en ESTRUCTURA (SIN DEDUPE) y US en hoja US (CON DEDUPE por DOC único).

    Retorna (nueva_fila_estructura, nueva_fila_us).
    """
    carpeta_csv = Path(carpeta_csv)

    # US dedupe persistente
    seen_u = _load_us_set(ws_control)

    # punteros de fila
    if fila_estructura is None:
        fila_estructura = _siguiente_fila_por_columna(ws_estructura, "E", fila_min=2)
    else:
        fila_estructura = max(int(fila_estructura), 2)

    if fila_us is None:
        fila_us = _siguiente_fila_por_columna(ws_us, "B", fila_min=2)
    else:
        fila_us = max(int(fila_us), 2)

    # ---- AT / AP / AC -> ESTRUCTURA (SIN DEDUPE)
    cell = ws_estructura.cell  # micro-opt

    for tipo in ("AT", "AP", "AC"):
        csv_path = next(carpeta_csv.glob(f"{tipo}*.CSV"), None)
        if not csv_path:
            print(f"[CARGA] {tipo} no encontrado")
            continue

        print(f"[CARGA] Procesando {tipo} | archivo={csv_path.name}")
        pairs = _MAPEO_IDX[tipo]

        count = 0
        for row in iterar_csv_sin_header(csv_path):
            for csv_i, col_idx in pairs:
                if csv_i < len(row):
                    cell(row=fila_estructura, column=col_idx, value=row[csv_i])
            fila_estructura += 1
            count += 1

            if imprimir_cada and (count % imprimir_cada == 0):
                print(f"[CARGA] {tipo}: {count} filas | fila_excel={fila_estructura}")

        print(f"[CARGA] Finalizado {tipo} ({count} filas)")

    # ---- US -> hoja US, col B, DOC único
    csv_us = next(carpeta_csv.glob("US*.CSV"), None)
    if csv_us:
        print(f"[CARGA] Procesando US | archivo={csv_us.name} | inicio_fila={fila_us}")

        escritos = 0
        saltados = 0
        leidos = 0

        cell_us = ws_us.cell

        for row in iterar_csv_sin_header(csv_us):
            leidos += 1
            if len(row) <= 1:
                continue

            doc = _norm_doc(row[1])
            if not doc:
                continue

            if doc in seen_u:
                saltados += 1
                continue

            cell_us(row=fila_us, column=2, value=doc)  # B=2
            fila_us += 1
            escritos += 1

            seen_u.add(doc)
            _append_control(ws_control, "U", doc)

            if imprimir_cada and (escritos % imprimir_cada == 0):
                print(f"[CARGA] US: escritos={escritos} | fila_excel={fila_us}")

        print(f"[CARGA] Finalizado US | leídos={leidos} | escritos={escritos} | duplicados_saltados={saltados}")
    else:
        print("[CARGA] US no encontrado")

    return fila_estructura, fila_us


# Legacy: abre/guarda por llamada (main optimizado NO debe usar esto)
def cargar_desde_carpeta(carpeta_csv: Path, fila_inicio: int) -> int:
    print("[CARGA] Abriendo plantilla Excel (modo legacy)...")
    wb, ws, ws_us, ws_control = abrir_plantilla()

    fila_estructura, _ = cargar_en_hojas(
        carpeta_csv=carpeta_csv,
        ws_estructura=ws,
        ws_us=ws_us,
        ws_control=ws_control,
        fila_estructura=fila_inicio,
        fila_us=None,
        imprimir_cada=1000,
    )

    print("[CARGA] Guardando Excel (modo legacy)...")
    wb.save(PLANTILLA)
    print("[CARGA] Excel guardado correctamente (modo legacy)")
    return fila_estructura


def main():
    import sys
    carpeta = Path(sys.argv[1])
    fila = int(sys.argv[2])
    ultima = cargar_desde_carpeta(carpeta, fila)
    print(f"ULTIMA_FILA={ultima}")


if __name__ == "__main__":
    main()
