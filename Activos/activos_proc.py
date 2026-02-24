from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
from datetime import datetime, date
import csv
import json
import re
import unicodedata
import openpyxl
from excel_com import norm_doc

_re_spaces = re.compile(r"\s+")

def _strip_accents(s: str) -> str:
    return "".join(
        ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch)
    )

def norm_servicio(s) -> str:
    if s is None:
        return ""
    s = str(s)
    s = _strip_accents(s)
    s = s.upper().strip()
    s = _re_spaces.sub(" ", s)
    return s

def parse_fecha_usuario(s: str) -> date:
    s = (s or "").strip()
    if not s:
        raise ValueError("Fecha vacía")
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    raise ValueError("Formato inválido. Use YYYY-MM-DD")

@dataclass
class ActivoRow:
    rownum: int
    tipo_doc: str
    doc_raw: str
    doc_norm: str
    servicio_raw: str
    servicio_norm: str

@dataclass
class PlanRow:
    tipo_doc: str
    doc_norm: str
    fecha: date
    codigo: str
    nombre_homologado: str
    l_base: str
    m_base: str
    base_row: int
    servicio_raw: str

def leer_activos_xlsx(xlsx_path: Path, sheet_name: str = "DETALLADO") -> list[ActivoRow]:
    print(f"    ... Cargando libro: {xlsx_path.name}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No existe hoja '{sheet_name}'")
    ws = wb[sheet_name]
    
    max_row = ws.max_row
    print(f"    ... Procesando {max_row} filas...")

    rows = []
    for r in range(2, max_row + 1):
        if r % 1000 == 0:
            print(f"        -> Leídas {r} filas...", end="\r")
            
        doc = ws.cell(r, 2).value
        serv = ws.cell(r, 5).value
        
        if doc or serv:
            rows.append(
                ActivoRow(
                    rownum=r,
                    tipo_doc="",
                    doc_raw=str(doc or "").strip(),
                    doc_norm=norm_doc(doc),
                    servicio_raw=str(serv or "").strip(),
                    servicio_norm=norm_servicio(serv)
                )
            )
    print(f"\n    ✅  Lectura completada. {len(rows)} filas válidas.")
    return rows

def cargar_mapeo_activos(json_path: Path) -> dict:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    return {
        norm_servicio(i["entrada"]): {
            "transformacion": (i.get("transformacion") or "").strip(),
            "codigo": (i.get("codigo") or "").strip()
        } 
        for i in data if norm_servicio(i.get("entrada"))
    }

def construir_plan_activos(excel, activos, mapeo, fecha):
    us_keys = excel.cargar_us_keyset()
    base_map = excel.cargar_estructura_base_lm()
    dupes = excel.cargar_estructura_dedupe_activos()
    
    doc_to_tipo = {}
    for k in us_keys:
        parts = k.split("|")
        if len(parts) == 2:
            doc_to_tipo[parts[1]] = parts[0]

    plan = []
    descartes = []
    fecha_iso = fecha.isoformat()

    for a in activos:
        if not a.doc_norm or not a.servicio_norm:
            descartes.append({"row_excel": a.rownum, "reason": "DOC_O_SERV_VACIO", "servicio": a.servicio_raw})
            continue

        tipo = doc_to_tipo.get(a.doc_norm)
        if not tipo:
            descartes.append({"row_excel": a.rownum, "reason": "NO_EXISTE_EN_US", "servicio": a.servicio_raw})
            continue

        m = mapeo.get(a.servicio_norm)
        if not m:
            descartes.append({"row_excel": a.rownum, "reason": "NO_MAPEO_EN_JSON", "servicio": a.servicio_raw})
            continue
            
        nombre_h = m.get("transformacion")
        if nombre_h == "QUITAR":
             descartes.append({"row_excel": a.rownum, "reason": "SERVICIO_EXCLUIDO", "servicio": a.servicio_raw})
             continue

        base = base_map.get(a.doc_norm)
        if not base:
             descartes.append({"row_excel": a.rownum, "reason": "NO_BASE_ESTRUCTURA", "servicio": a.servicio_raw})
             continue
        
        l_val = str(base["L"]).strip()
        m_val = str(base["M"]).strip()
        
        if not l_val or not m_val:
            descartes.append({"row_excel": a.rownum, "reason": "BASE_SIN_LM", "servicio": a.servicio_raw})
            continue

        codigo = m["codigo"]
        key_dupe = f"{a.doc_norm}|{codigo}|{fecha_iso}"
        if key_dupe in dupes:
            descartes.append({"row_excel": a.rownum, "reason": "DUPLICADO_YA_EXISTE", "servicio": a.servicio_raw})
            continue

        plan.append(
            PlanRow(
                tipo_doc=tipo,
                doc_norm=a.doc_norm,
                fecha=fecha,
                codigo=codigo,
                nombre_homologado=nombre_h,
                l_base=l_val,
                m_base=m_val,
                base_row=base["row"],
                servicio_raw=a.servicio_raw
            )
        )
        
    return plan, descartes

def exportar_auditoria_csv(path, plan, descartes):
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["TIPO", "tipo_doc", "doc", "fecha", "codigo", "nombre_homologado", "L", "M", "base_row", "servicio", "reason", "extra", "row_excel"])
        
        for p in plan:
            w.writerow(["OK", p.tipo_doc, p.doc_norm, p.fecha, p.codigo, p.nombre_homologado, p.l_base, p.m_base, p.base_row, p.servicio_raw, "", "", ""])
            
        for d in descartes:
            w.writerow(["NO", "", "", "", "", "", "", "", "", d.get("servicio"), d.get("reason"), "", d.get("row_excel")])